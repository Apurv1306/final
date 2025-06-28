import cv2
import os
import numpy as np
import time
from playsound import playsound
from datetime import datetime
from openpyxl import Workbook, load_workbook

# === Paths and Config ===
KNOWN_FACES_DIR = "known_faces"
AUDIO_FILE = "thank_you.mp3"
TICK_ICON_PATH = "tick.png"
EXCEL_FILE = "Attendance.xlsx"
RECOGNITION_INTERVAL = 600  # 10 minutes

# === Ensure Excel file exists ===
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    ws.append(["Date", "Employee ID", "Name", "Timestamp"])
    wb.save(EXCEL_FILE)

# === Load tick image ===
tick_img = cv2.imread(TICK_ICON_PATH, cv2.IMREAD_UNCHANGED)
if tick_img is None:
    raise FileNotFoundError("tick.png not found in project folder.")

tick_img = cv2.resize(tick_img, (30, 30))  # Resize icon

# === Load Haar Cascade ===
face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml")

# === Prepare training data ===
faces = []
labels = []
label_map = {}
name_to_label = {}
label_counter = 0
employee_ids = {}

for file in os.listdir(KNOWN_FACES_DIR):
    if file.endswith(".jpg") or file.endswith(".png"):
        parts = file.split("_")
        if len(parts) >= 2:
            name = parts[0].capitalize()
            emp_id = parts[1].split(".")[0]
            path = os.path.join(KNOWN_FACES_DIR, file)
            img = cv2.imread(path, cv2.IMREAD_GRAYSCALE)

            if img is None:
                continue

            img = cv2.resize(img, (200, 200))

            if name not in name_to_label:
                name_to_label[name] = label_counter
                label_map[label_counter] = name
                employee_ids[label_counter] = emp_id
                label_counter += 1

            faces.append(img)
            labels.append(name_to_label[name])

labels = np.array(labels)

recognizer = cv2.face.LBPHFaceRecognizer_create()
recognizer.train(faces, labels)

cap = cv2.VideoCapture(0)
last_recognition_time = {}

def overlay_image(background, overlay, x, y):
    h, w = overlay.shape[:2]
    if x + w > background.shape[1] or y + h > background.shape[0]:
        return background
    for i in range(h):
        for j in range(w):
            if overlay[i, j][3] != 0:
                background[y+i, x+j] = overlay[i, j][:3]
    return background

def log_attendance(emp_id, name):
    now = datetime.now()
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([now.strftime("%Y-%m-%d"), emp_id, name, now.strftime("%H:%M:%S")])
    wb.save(EXCEL_FILE)

while True:
    ret, frame = cap.read()
    if not ret:
        break

    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    detected_faces = face_cascade.detectMultiScale(gray, 1.3, 5)

    for (x, y, w, h) in detected_faces:
        roi_gray = gray[y:y+h, x:x+w]
        roi_resized = cv2.resize(roi_gray, (200, 200))

        label, confidence = recognizer.predict(roi_resized)

        if confidence < 70:
            name = label_map.get(label, "Unknown")
            emp_id = employee_ids.get(label, "N/A")
            now = time.time()
            last_time = last_recognition_time.get(label, 0)

            cv2.putText(frame, f"{name}", (x, y - 10),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 255, 0), 2)
            current_time_str = datetime.now().strftime("%H:%M:%S")
            cv2.putText(frame, current_time_str, (x, y + h + 25),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)

            if now - last_time > RECOGNITION_INTERVAL:
                playsound(AUDIO_FILE)
                log_attendance(emp_id, name)
                last_recognition_time[label] = now

            cv2.rectangle(frame, (x, y), (x+w, y+h), (255, 0, 0), 2)
            frame = overlay_image(frame, tick_img, x + w - 35, y + 5)
        else:
            cv2.putText(frame, "Unknown", (x, y - 10),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 0, 255), 2)
            cv2.rectangle(frame, (x, y), (x+w, y+h), (0, 0, 255), 2)

    cv2.imshow("Face Recognition", frame)
    key = cv2.waitKey(1)
    if key == ord("q"):
        break

cap.release()
cv2.destroyAllWindows()
